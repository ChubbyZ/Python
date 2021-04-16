# 新增功能：函数化，cmd窗口传参,路透社
# encoding='utf-8
import argparse
import json
import re

import feedparser
import pprint

import requests
from chardet import detect
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import urllib3
import ssl
urllib3.disable_warnings()
# ssl._create_default_https_context = ssl._create_unverified_context()
a = 0
mylists = []

proxy = '127.0.0.1:10809'
proxies = {
    'http': 'http://' + proxy,
    'https': 'http://' + proxy
}


def arg_parse():
    parser = argparse.ArgumentParser(usage='python %(prog)s --word')
    parser.add_argument('-w', type=str, metavar='关键词', help='要筛选信息关键词')
    parser.add_argument('-f', type=str, metavar='要打开或存储的文件', help='要执行的命令')
    args = parser.parse_args()
    return args


# 网站种子解析,遍历，挨个解析
def get_rss_oschina():
    with open("rss.txt", 'r+') as fr:
        for url in fr:
            url = url.split()
            rss_oschina = feedparser.parse(url[0])
            get_link(rss_oschina)
    return rss_oschina


def get_link(rss_oschina):
    channel_lists = []
    link_lists = []
    # 整理为JSON数组
    mylist = [{'title': entry['title'], 'link': entry['link']} for entry in rss_oschina['entries']]
    for i in range(len(mylist)):
        channel_lists.append(mylist[i]['title'])
        link_lists.append(mylist[i]['link'])
    write_data(channel_lists, link_lists, get_feed_title(rss_oschina))


def get_feed_title(rss_oschina):
    global title
    if rss_oschina.feed.__contains__('title'):
        title = rss_oschina.feed.title
        pprint.pprint(title)
    else:
        pass
    return title


def write_data(channel_lists, link_lists, title):
    # 判断系统是是不是有该文件，如果没有的话，自动创建，有的话，打开已存在文件进行写入，写入方式是追加
    if os.path.exists(file):
        wb = load_workbook(file)
    else:
        wb = Workbook()
    ws = wb.active
    # 将表格中已经存在的数据写入到列表中
    lists = []
    title_lists = []
    for row_title in ws.iter_cols(1, 1, max_row=ws.max_row):
        for cell_title in row_title:
            title_lists.append(cell_title.value)
    for row in ws.iter_cols(2, 2, max_row=ws.max_row):
        for cell in row:
            lists.append(cell.value)
    # 只把链接中和标题中含有solarwinds的写入到文件中
    for i in range(len(channel_lists)):
        if (f'{word}' in channel_lists[i]) or (f'{word}' in link_lists[i]):
            # 按行写入
            for row in range((ws.max_row + 1), (ws.max_row + len(channel_lists))):
                data = link_lists[i]
                # 去重
                if data not in lists:
                    ws.cell(row, 2, data)
                    if title not in title_lists:
                        ws.cell(row, 1, title)
                break
            continue
    wb.save(file)


class site_link_gather(object):
    reuters_links = []

    def reuters(self):

        reuters_links = []
        # try:
        global reuters_response
        # session = requests.session()
        # requests.DEFAULT_RETRIES = 5  # 增加重连次数
        # reuters_response = requests.session()
        # session.keep_alive = False  # 关闭多余连接
        # reuters_response = session.get(f'https://www.reuters.com/search/news?blob=solarwinds&sortBy=date&dateRange=all')

        reuters_response = requests.get(
            f'https://www.reuters.com/search/news?blob=solarwinds&sortBy=date&dateRange=all', proxies=proxies,
            verify=False)
        print(reuters_response.status_code)
        # except Exception as e:
        #     print(e)
        pat = '<a href="/article/([^\s]*)">'
        reuters_rst = re.findall(pat, reuters_response.text)
        # return reuters_rst
        for r in range(len(reuters_rst)):
            reuters_link = f'https://www.reuters.com/article/{reuters_rst[r]}'
            reuters_links.append(reuters_link)
        return reuters_links

    def cnn(self):
        global cnn_response, cnn_link
        try:
            cnn_response = requests.get(f'https://search.api.cnn.io/content?q={word}&sort=newest&size=20', verify=False,
                                        proxies=proxies)
        except Exception as e:
            print(e)
        cnn_pat = '"url":"([^\s]*)","firstPublishDate"'
        cnn_rst = re.findall(cnn_pat, cnn_response.text)
        return cnn_rst
        # for r in range(len(cnn_rst)):
        #     cnn_link = f'{cnn_rst[r]}'
        #     print(cnn_link)

    def washingtonpost(self):
        global response, reuters_link
        try:
            response = requests.get(
                f'https://sitesearchapp.washingtonpost.com/sitesearch-api/v2/search.json?count=20&datefilter=displaydatetime:%5B*+TO+NOW%2FDAY%2B1DAY%5D&facets.fields=%7B!ex%3Dinclude%7Dcontenttype,%7B!ex%3Dinclude%7Dname&highlight.fields=headline,body&highlight.on=true&highlight.snippets=1&query={word}&sort=displaydatetime+desc&callback=angular.callbacks._0',
                verify=False, proxies=proxies)
            print(response.status_code)
        except Exception as e:
            print(e)
        pat = '"contenttype":"Article","contenturl":"([^\s]*)","headline"'
        rst_washingtonpost = re.findall(pat, response.text)
        return rst_washingtonpost
        # for r in range(len(rst)):
        #     washingtonpost_link = f'{rst[r]}'
        #     print(washingtonpost_link)

    def write_data_site(self):
        if os.path.exists(file):
            wb = load_workbook(file)
        else:
            wb = Workbook()
        ws = wb.active
        # 将表格中已经存在的数据写入到列表中
        lists = []
        title_lists = []
        for row_title in ws.iter_cols(1, 1, max_row=ws.max_row):
            for cell_title in row_title:
                title_lists.append(cell_title.value)
        for row in ws.iter_cols(2, 2, max_row=ws.max_row):
            for cell in row:
                lists.append(cell.value)
        print('start')
        site_links = self.reuters() + self.cnn() + self.washingtonpost()
        print('end')
        i = 0
        print(ws.max_row)
        for insert in range((ws.max_row + 1), (ws.max_row + len(site_links))):
            if site_links[i] not in lists:
                ws.cell(row=ws.max_row + 1, column=2, value=site_links[i])
            i = i + 1
        wb.save(file)


if __name__ == '__main__':
    args = arg_parse()
    # word = args.w
    word = 'solarwinds'
    file = '收集链接2.xlsx'
    # file = args.f
    # get_rss_oschina()
    site_link_gather().write_data_site()
    get_rss_oschina()
